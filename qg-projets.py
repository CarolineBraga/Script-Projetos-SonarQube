import json
import logging
import sys

import gitlab
import requests
from openpyxl import load_workbook


def format_column(ws):
    for letter in "ABCDEFGHI":
        ws.column_dimensions[letter].width = 60


def fetch_qg_data(wb):
    """Fetch SonarQube projects data and write to quality gate worksheet"""
    ws = wb["Quality Gate"]
    format_column(ws)

    response = requests.get(
        "http://sonarqubeexample.com/api/projects/search",
        auth=(sys.argv[1], ""),
    )
    sonarqube = json.loads(response.text)

    row = 2
    col = 9
    qg_projects = set()

    for p in sonarqube["components"]:
        pos = p["key"].find(":")
        qg_projects.add(p["key"][0:pos])

    for p in qg_projects:
        row += 1
        ws.cell(row, col, p)

    row -= 1
    return qg_projects


def fetch_gl_data(wb, qg):
    """Fetch GitLab data and write to gitlab worksheet"""
    projects = {}
    ws = wb["GitLab"]

    format_column(ws)

    gl = gitlab.Gitlab("https://gitlab.com/", private_token=sys.argv[2])
    gl.auth()

    domains_id = {
        "Arquitetura": 56559265,
        "B2B": 54189372,
        "B2C": 54191218,
        "Dados": 56559289,
        "Industria": 56559298,
        "Operacoes": 54510913,
        "Seguranca": 56559308,
        "SC": 56559315,
    }

    domains_qg = {
        "Arquitetura": 0,
        "B2B": 0,
        "B2C": 0,
        "Dados": 0,
        "Industria": 0,
        "Operacoes": 0,
        "Seguranca": 0,
        "SC": 0,
    }

    col = 1

    for domain in domains_id:
        group = gl.groups.get(domains_id[domain])
        gl_projects = group.projects.list(include_subgroups=True, get_all=True)

        row = 2
        qg_row = 3

        for p in gl_projects:
            path = p.attributes["path_with_namespace"]
            row += 1
            pos = path.rfind("/")
            name = path[pos + 1 :]
            ws.cell(row, col, name)
            if name in qg:
                ws = wb["Quality Gate"]
                ws.cell(qg_row, col, name)
                domains_qg[domain] = domains_qg[domain] + 1
                qg_row += 1
                ws = wb["GitLab"]

        projects[domain] = row - 1
        col += 1

    return [projects, domains_qg]


def put_numbers(wb, projects_qty):
    """Write the numbers to Numeros worksheet"""
    ws = wb["Números"]

    row = 2
    for np in projects_qty[0].values():
        ws.cell(row, 2, np)
        row += 1

    row = 2
    for nqg in projects_qty[1].values():
        ws.cell(row, 3, nqg)
        row += 1

    ws["B10"] = "=SUM(B2:B9)"
    ws["C10"] = "=SUM(C2:C9)"

    for r in range(2, 11):
        row = str(r)
        ws["D" + row] = "=(C" + row + "/B" + row + ")"


def fetch_issues(wb):
    """Fetch SonarQube issues numbers and write to Numeros worksheet"""
    ws = wb["Números"]
    col = 5

    response = requests.get(
        "http://35.199.100.34:9000/api/issues/search?ps=1&facets=types",
        auth=("squ_007cdd9b009e86fe7922a13af2f7be64aa3d91c0", ""),
    )
    sonarqube = json.loads(response.text)
    issues = sonarqube["facets"][0]["values"]

    i = 0
    while i < len(issues):
        ws.cell(1, col, issues[i]["val"])
        ws.cell(10, col, issues[i]["count"])
        i += 1
        col += 1


def save_workbook(wb):
    """Save updated repositorios.xlsx"""
    wb.save("repositorios.xlsx")


def get_workbook():
    """Load repositorios.xlsx workbook"""
    try:
        wb = load_workbook("./repositorios.xlsx")
        qg = fetch_qg_data(wb)
        projects_qty = fetch_gl_data(wb, qg)
        put_numbers(wb, projects_qty)
        fetch_issues(wb)
        save_workbook(wb)

    except FileNotFoundError:
        logging.error("File does not exist.")


def main():
    get_workbook()


if __name__ == "__main__":
    main()
